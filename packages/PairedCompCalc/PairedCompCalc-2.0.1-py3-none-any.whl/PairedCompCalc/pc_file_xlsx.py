"""Read paired-comparison data stored in an Excel 2007- (xlsx) workbook file.

Each Excel work-book file, and each work-sheet in the file,
may include paired-comparison results for
one or more subject(s), tested in one or more test-factor combinations,
for one or more perceptual attributes.

Each row in the worksheet must include enough data to specify
one stimulus pair and the corresponding (possibly graded) judgment response,
to be represented by ONE PairedCompItem instance.

The perceived difference between the two sounds in a presented pair (stim0, stim1)
may be encoded in three different ways:
*1: a single signed integer in (-M, ..., +M) showing the signed difference,
    where a positive value means that stim1 was the "better" (preferred) sound in a pair (stim0, stim1),
*2: a combination of (choice, grade), where
    choice is a string == either stim0 or stim1,
    or a string in the parameter no_choice, indicating no perceived difference.
    The grade is a string identifying the magnitude of the difference.
    The grade must then be an element in the list PairedCompFrame.difference_grades.
*3: only the string choice, if the judgment is binary, i.e., with no difference grading.

A judgment of "no difference", i.e., numeric difference = 0, is allowed
only if the PairCompFrame instance has property forced_choice == False.
This judgement can then be represented EITHER by
choice == an element in the list no_choice, without any defined magnitude grade,
OR integer-coded difference == 0.


*** Usage Example:

pcf=PairedCompFrame(objects=['HA-1', 'HA-2', 'HA-3'],
            attributes=['Preference'],
            difference_grades=['equal', 'slightly better', 'better', 'much better'],
            test_factors=dict(Sound=['speech', 'music'], SNR=['low', 'high']),
            forced_choice=False
            )

pc_file = PairedCompFile(file_path,
        pcf=pcf,
        sheets=[f'Subject{i} for i in range(10)],
        subject='sheet',
        top_row=4,  # first row with paired-comparison data
        attribute='A',
        pair=('B', 'C'),
        difference='D',  # OR choice='E' and grade='F', no_choice=(None, '', 'None')
        Sound='G'  # category for Test Factor Sound, as defined in pcf.test_factors
        )

for pc_item in pc_file:
    process the pc_item

The parameter pcf defines experimental layout and
possible categories within each given test factor to be analyzed.

The parameter sheets is a list of workbook sheets to be searched for results.

Other PairedCompFile properties define locations for
subject, attribute, stimulus-pair labels, difference magnitude,
and test factors.

The location of subject ID, attribute, and a test factor can be
EITHER a column address, like 'C',
OR 'sheet', indicating that the sheet name is to be used for all included items.
Obviously, only ONE of these properties can be stored in the sheet name.

In this example, subject='sheet' indicates that the sheet name is interpreted as subject id.
pair=('B', 'C') defines columns containing stimulus labels.
This parameter may be omitted if pcf.objects has exactly TWO elements,
and then it is assumed that every pair == pcf.objects.

attribute='A' defines the column defining the perceptual-attribute label.
This parameter may be omitted if pcf.attributes includes exactly ONE element,
as in this example.

Thus, in this example, each row x, for x >= 4, specifies
the pair (stim0, stim1) in cells Bx and Cx,
the integer-coded difference in cell Dx,
and cell Gx specifies a category for test-factor 'Sound'.

The file path string will be examined to find
one of the allowed categories for the remaining test-factor 'SNR',
which does not have a column address.


*** Version History:
* Version 2.0.1:
2022-03-08, minor fix for more informative error messag

* Version 2.0:
2021-09-18, changed signature for PairedCompFile.__init__(). New argument grade!
2021-09-18, new method save
2021-09-19, allow only signed-integer-encoded response in field 'difference',
            OR label-coded as ('choice', 'grade') fields.
            Previous version allowed 'difference' to also mean 'grade' category
2021-09-22, revised method save, tested, changed signature

*Version 1.0:
2018-10-01, first functional version
"""
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.cell import column_index_from_string
import logging

from PairedCompCalc import pc_base

logger = logging.getLogger(__name__)
# logger.setLevel(logging.DEBUG)  # TEST


class FileFormatError(pc_base.FileReadError):
    """Format error causing non-usable data"""


class ParameterError(pc_base.FileReadError):
    """Error in calling parameters causing non-usable data in some file."""


class ArgumentError(RuntimeError):
    """Error in calling arguments. No file can be read."""


class PairedCompFile(pc_base.PairedCompFile):
    """Interface to Excel xlsx workbook file storing paired-comparison data.
    Each sheet in the workbook may include response data for
    one or more subjects, attributes, and test conditions.
    Data elements MUST be stored at the same location in all sheets.
    """
    def __init__(self, file_path, pcf,
                 subject,
                 attribute=None,
                 pair=None,
                 difference=None,
                 choice=None,
                 grade=None,
                 no_choice=(None, '', 'None'),
                 top_row=2,
                 sheets=None,
                 **test_factors):
        """File interface to data stored in an excel file,
        decoded as a sequence of PairedCompItem instances.
        :param file_path: Path to file for reading
        :param pcf: single PairedCompFrame instance
        :param subject: 'sheet' or column for subject identification label
        :param attribute: (optional) 'sheet' or column for perceptual attribute labels
            May be None if len(pcf.attributes) == 1
        :param pair: (optional) tuple with two column addresses defining (stim0, stim1)
            may be None if there are only two objects in total for comparison
        :param difference: (optional) column with perceived difference, coded as
            an INTEGER indicating the ranked difference of stim1 rel. stim0,
            with positive value indicating that stim1 was better than stim0
        :param choice: (optional) column with STRING label of preferred pair element.
            May be None, only if difference is not None.
        :param grade: (optional) column with perceived difference, coded as
            a label == one of pcf.difference_grades
            May be None if difference is Not None.
            May be None if only one non-equal grade is allowed,
            i.e., difference can be only -1, 0, or +1.
        :param no_choice: (optional) sequence of possible cell values in the 'choice' column,
            indicating no audible difference or no preference.
        :param top_row: integer address of first row containing PairedCompItem data
        :param sheets: (optional) list of sheet names to be searched for data
            if None, all sheets are searched.
        :param test_factors: (optional) dict with elements (tf, location), where
            tf is a string identifying one test factor,
            location is a string with a column address like 'D'.
            For file reading: Any unspecified test factor must instead be obtained from file path.
            For file writing: All pcf.test_factors.keys() must have specified column address.
        """
        super().__init__(file_path, pcf)
        self.subject = _check_column_or_sheet(subject)
        self.attribute = _check_column_or_sheet(attribute)
        try:
            self.pair = (_check_column(pair[0]), _check_column(pair[1]))
        except (TypeError, IndexError):
            self.pair = None
        self.choice = _check_column(choice)
        self.no_choice = no_choice
        self.difference = _check_column(difference)
        if (self.difference is None) and (self.choice is None):
            raise ArgumentError('Either "difference" or "choice" must have defined address')
        if (grade is None) and (self.choice is not None) and (self.difference is not None):
            self.grade = self.difference  # *** only for backward compatibility  ***
            self.difference = None  # ********** using grade instead
            logger.warning('Using column "difference" only for integer-coded responses.\n'
                           + f'Changed to "grade" for ordinal labels in {file_path}.')
        else:
            self.grade = _check_column(grade)
        self.top_row = top_row
        self.sheets = sheets
        self.test_factors = _check_test_cond(test_factors)
        # **** following used only by save method:
        field_address = {'subject': self.subject,
                         'attribute': self.attribute,
                         'pair_0': None if self.pair is None else self.pair[0],
                         'pair_1': None if self.pair is None else self.pair[1],
                         'difference': self.difference,
                         'choice': self.choice,
                         'grade': self.grade,
                         }
        field_address.update(self.test_factors)
        self.field_index = {k: (None if (v == 'sheet') or (v is None)
                                else column_index_from_string(v) - 1)
                            for (k, v) in field_address.items()}
        self.sheet_field = None  # only ONE field may be encoded in sheet title
        for field in field_address.keys():
            if field_address[field] == 'sheet':
                if self.sheet_field is None:
                    self.sheet_field = field
                else:
                    raise ArgumentError('Only ONE property can be stored in "sheet" title')

    def __iter__(self):
        """Generator yielding data from an excel file.
        :return: generator yielding PairedCompItem instance
        """
        try:
            wb = load_workbook(str(self.file_path), read_only=True)
        except Exception as e:  # {InvalidFileException, BadZipFile}:
            raise FileFormatError(f'Cannot load Excel workbook from file {self.file_path.stem}. '
                                  + f'Openpyxl error: {str(e)}')
        if self.sheets is None:
            sheets = wb.sheetnames
        else:
            sheets = set(self.sheets) & set(wb.sheetnames)
        if len(sheets) == 0:
            raise FileFormatError(f'No accepted sheets found in {self.file_path}')
        path_test_cond = self.path_test_cond()
        for sheet_name in sheets:
            ws = wb[sheet_name]
            logger.debug(f'Sheet {ws.title} in {self.file_path} has max_row= {ws.max_row}')
            rows = ws.rows
            for _ in range(self.top_row - 1):
                try:
                    row = rows.__next__()
                    logger.debug(f'skipping row {row[0].row}')
                except StopIteration:
                    raise FileFormatError(f'No data rows in {self.file_path.stem}.'
                                          + f'sheet {repr(sheet_name)}')
            for row in rows:
                if all(c.value is None for c in row):
                    logger.debug(f'Empty row encountered. Stop reading file {self.file_path}')
                    break  # end of valid part of file
                logger.debug(f'Processing row {row[0].row}')
                p = self._get_pair(row)
                r = pc_base.PairedCompItem(subject=self._get_subject(ws, row),
                                           attribute=self._get_attribute(ws, row),
                                           pair=p,
                                           response=self._get_response(row, p),
                                           test_cond=self._get_test_cond(ws, row, path_test_cond))
                if not any(r_i is None for r_i in r.__dict__.values()):
                    # we have a valid item
                    yield r
                else:
                    logger.debug(f'Incomplete item from row {row[0].row} in file {self.file_path}:\n'
                                 + repr(r))

    def _get_subject(self, ws, row):
        """
        :param ws: a worksheet
        :param row: tuple of openpyxl Cell instances
        :return: subject code, integer or string
        """
        return _get_sheet_or_cell_value(ws, self.subject, row)

    def _get_attribute(self, ws, row):
        """
        :param ws: a worksheet
        :param row: tuple of openpyxl Cell instances
        :return: string label
        """
        if self.attribute is None:
            if len(self.pcf.attributes) == 1:
                return self.pcf.attributes[0]
            else:
                raise ArgumentError(f'Undefined attribute in file {self.file_path.stem}')
        else:
            return _get_sheet_or_cell_value(ws, self.attribute, row)

    def _get_pair(self, row):
        """
        :param row: tuple of openpyxl Cell instances
        :return: tuple of strings (stim0, stim1)
        """
        if self.pair is None:
            if len(self.pcf.objects) == 2:
                return tuple(self.pcf.objects)
            else:
                raise ArgumentError(f'Undefined pair address')
        stim0 = _get_cell_value(self.pair[0], row)
        stim1 = _get_cell_value(self.pair[1], row)
        if (stim0 is None) or (stim1 is None):
            return None
        else:
            return stim0, stim1

    def _get_response(self, row, pair):
        """Encode response as an integer in (-M, ..., +M)
        Responses may be coded either as (choice, grade)
        OR simply as difference = signed integer, with sign indicating the choice
        :param row: tuple of openpyxl Cell instances
        :param pair: tuple (stim0, stim1), either default or from this row
        :return: integer in (-M,...,M) indicating ranked difference

        2021-09-19, cleanup, allowing only cell input
        2021-09-19, bug fix for grade magnitude in case forced_choice
        """
        if self.difference is not None:  # using column with integer-coded response
            diff = _get_cell_value(self.difference, row)
            if type(diff) is int:  # signed integer
                return diff
            elif diff is None:
                logger.debug(f'Row with "difference" = {repr(diff)} disregarded.')
                return None  # incomplete row, skip decoding
            else:  # try string label interpreted as 'grade' instead
                logger.debug(f'Non-integer "difference" = {repr(diff)} disregarded. '
                             + 'Trying ("choice","grade") instead.')
        # must instead decode string-coded (choice, grade) as difference:
        if self.choice is None:
            raise ArgumentError('No choice address defined')
        else:
            choice = _get_cell_value(self.choice, row)
        if choice in self.no_choice:
            if self.pcf.forced_choice:
                logger.warning(f'choice -> no_choice in file {self.file_path}, '
                               + 'although forced_choice. Should not happen!')
            return 0  # use as valid response anyway
        if pair is None:
            logger.debug(f'Undefined pair: Cannot decode choice')
            return None  # error: need pair to determine response_sign
        if choice == pair[0]:
            response_sign = -1
        elif choice == pair[1]:
            response_sign = 1
        else:
            logger.debug(f'choice={repr(choice)} does not match pair {pair}')
            return None  # should not happen. Error!
        if self.grade is None:  # only binary choice, no grading
            return response_sign
        else:
            response_grade = _get_cell_value(self.grade, row)
            try:
                response_magn = self.pcf.difference_grades.index(response_grade)
                if self.pcf.forced_choice:  # grade Equal not allowed
                    response_magn += 1  # difference_grades[0] means 1
                return response_sign * response_magn
            except ValueError:  # Non-existent grade
                logger.debug(f'Undefined grade = {repr(response_grade)} ')
                return None

    def _get_test_cond(self, ws, row, path_tc):
        """Get tuple of test conditions
        :param ws: a worksheet
        :param row: tuple of openpyxl Cell instances
        :param path_tc: default dict with (tf, tf_category) from file path
        :return: dict with elements (tf, tf_category)
        """
        tc = path_tc.copy()
        for (tf, col) in self.test_factors.items():
            tc[tf] = _get_sheet_or_cell_value(ws, col, row)
        return tc

    # ----------------------------------------------- save data to workbook
    def save(self, items, allow_over_write=False):
        """Save self.items to file.
        Usually, all item properties are stored in separate column,
        but just like read-only files,
        ONE of properties subject, attribute, or one test factor,
        may be stored in the sheet.title instead.
        :param items: iterable of PairedCompItem instances
        :param allow_over_write: boolean switch, =True allows old file to be over-written
        :return: None

        Result: a new file is created.

        2021-09-16, new method for this file format
        """
        # *** update self properties to allow reading from the file?
        wb = Workbook(write_only=True)
        n_col = 1 + max(v for v in self.field_index.values() if v is not None)
        for field in ['difference', 'choice', 'grade'] + [*self.pcf.test_factors.keys()]:
            # *** include all for writing, although not all mandatory for reading ?
            if field not in self.field_index.keys() or self.field_index[field] is None:
                self.field_index[field] = n_col
                n_col += 1
        header = self._make_header(n_col)
        self.top_row = 2
        for pci in items:
            ws = self._get_sheet(wb, pci, header)
            ws.append(self._make_row(pci, n_col))
        if not allow_over_write:
            self.file_path = pc_base.safe_file_path(self.file_path)
        self.sheets = wb.sheetnames
        wb.save(self.file_path)

    def _get_sheet(self, wb, pci, header):
        """Select or create sheet for saving
        :param wb: current workbook instance
        :return: an openpyxl.WorkSheet instance with correct title and filled header
        """
        if self.sheet_field is None:
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet()
                ws.append(header)
            return ws
        elif self.sheet_field == 'subject':
            sh_title = pci.subject
        elif self.sheet_field == 'attribute':
            sh_title = pci.attribute
        else:
            raise ArgumentError('Only "subject" or "attribute" may be saved as "sheet" title')
        if sh_title in wb.sheetnames:  # this sheet has been created and used before
            ws = wb[sh_title]
        else:
            ws = wb.create_sheet(title=sh_title)
            ws.append(header)
        return ws

    def _make_header(self, n_col):
        """Make header list from column addresses of self.
        :param n_col:
        :return: list of header labels, length = n_col
        """
        h = [None for _ in range(n_col)]
        for (k, v) in self.field_index.items():
            if v is not None:
                h[v] = k
        return h

    def _make_row(self, pci, n_col):
        """Make list with contents from given paired-comp data instance
        :param pci: a pc_base.PairedCompItem instance
        :param n_col: length of row
        :return: r = list filled with elements from pci,
            placed at column indices in self.field_index.
        """
        # always save response BOTH as
        # int-coded difference AND string-coded as (choice, grade)
        # to exemplify both coding methods, even if not requested by user
        diff_int = pci.response
        if self.pcf.forced_choice:
            diff_grade = self.pcf.difference_grades[abs(diff_int) - 1]
        else:
            diff_grade = self.pcf.difference_grades[abs(diff_int)]
        if diff_int > 0:
            choice = pci.pair[1]
        elif diff_int < 0:
            choice = pci.pair[0]
        else:
            choice = 'None'  # 'None'
        r_dict = {'subject': pci.subject,
                  'attribute': pci.attribute,
                  'pair_0': pci.pair[0],
                  'pair_1': pci.pair[1],
                  'difference': diff_int,  # integer-coded
                  'choice': choice,
                  'grade': diff_grade
                  }
        r_dict.update(pci.test_cond)
        r = [None for _ in range(n_col)]
        for (k, v) in r_dict.items():
            if self.field_index[k] is not None:
                r[self.field_index[k]] = v
        return r


# --------------------------------------------------- help sub-functions

def _check_column(col):
    """Check that a column address is acceptable for openpyxl
    :param col: None, or string with column address
    :return: col, or None
    """
    if col is None:  # acceptable for some fields
        return col
    try:
        column_index_from_string(col)
        return col
    except (ValueError, AttributeError) as e:
        raise ArgumentError(f'Invalid column string address: {col}')


def _check_column_or_sheet(col):
    """Check that a parameter address is either 'sheet' or a column string address
    """
    if col == 'sheet':
        return col
    else:
        return _check_column(col)


def _check_test_cond(tf_address):
    """Check dict with column (or sheet) addresses
    :param tf_address: dict with elements (test_factor, address)
    :return: tf_address if OK, else None
    """
    if not any(_check_column_or_sheet(col) is None
               for col in tf_address.values()):
        return tf_address
    else:
        logger.warning(f'Incorrect address in {tf_address}')
        return None


def _get_cell_value(col, row):
    """Get one cell value from a row
    :param col: column address
    :param row: tuple of openpyxl.Cell instances
    :return: cell value, string or number
    """
    try:
        c = row[column_index_from_string(col) - 1].value
        if type(c) is str and '=' in c:
            logger.warning(f'Cell {col}{row[0].row}= {repr(c)} seems to be a formula. Not evaluated.')
        return c
        # return row[column_index_from_string(col) - 1].value
    except KeyError:
        return None


def _get_sheet_or_cell_value(ws, col, row):
    """Get contents in ONE cell or in sheet title
    :param ws: a worksheet
    :param col: one column address or 'sheet'
    :param row: integer address of row
    :return: cell contents
    """
    if col == 'sheet':
        return ws.title
    else:
        return _get_cell_value(col, row)
